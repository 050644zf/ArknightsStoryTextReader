import openpyxl as xl
import re
import argparse
from pathlib import Path
import os
from func import getFile
import func
from openpyxl.worksheet.hyperlink import Hyperlink
from jsonconvert import reader
import sys


characters = []
codes = ['--Branch--','----','--Decision End--',]
infoFlag = False
commentFlag = False

bold = xl.styles.Font(b=True)
underline = xl.styles.Font(u='single')


def xlc(sheet, storyJson):
    rawlist = storyJson['storyList']
    for idx, line in enumerate(rawlist):
        print(f"\rExporting Line {idx+1}/{len(rawlist)} ... ", end='')
        index = line['id']
        prop = line['prop']
        attrs = line['attributes']

        if not f'--{prop}--' in codes:
            codes.append(f'--{prop}--')

        if prop == 'Comment' and commentFlag:
            sheet.append([index,
            '--Comment--',
            attrs['value']])
        
        if prop == 'name':
            sheet.append([index,
            attrs.get('name',''),
            attrs['content']])
            if not attrs.get('name','') in characters:
                characters.append(attrs.get('name',''))

        if prop == 'multiline':
            sheet.append([
                index,
                attrs.get('name',''),
                attrs['content']
            ])
            if not attrs.get('name','') in characters:
                characters.append(attrs.get('name',''))

        if prop == 'Dialog':
            sheet.append([index,
            '----',
            '----'])

        if prop == 'Decision':
            sheet.append([index,
            '--Decision--',
            '----'])
            options = attrs['options'].split(';')
            values = attrs['values'].split(';')
            for i in range(min(len(values),len(options))):
                sheet.append([index,
                f'Option_{values[i]}',
                options[i]])
            sheet.append([index,
            '--Decision End--',
            '----'])
        
        if prop == 'Predicate':
            if line.get('endOfOpt'):
                sheet.append([index,
                '--Branch--',
                '> End of Options'])
            else:
                sheet.append([index,
                '--Branch--',
                f'> Options_{attrs["references"].replace(";","&")}'])

        if prop == 'Subtitle':
            sheet.append([''])
            sheet.append([index,
            '',
            attrs.get('text','')])
            sheet.append([''])

        if prop == 'Sticker':
            sheet.append([''])
            sheet.append([index,
            '',
            attrs.get('text','')])

        if prop == 'stickerclear':
            sheet.append([''])

        

        if 'image' in attrs.keys():
            prop = prop.lower()
            if prop == "backgroundtween":
                prop = "background"
            elif prop == 'showitem':
                prop = 'item'
            
            sheet.append([index,
            f'--{prop}--',
            f'https://aceship.github.io/AN-EN-Tags/img/avg/{prop}s/{attrs["image"]}.png'])

        if not f'--{prop}--' in codes:
            codes.append(f'--{prop}--')
        
if __name__ == "__main__":

    os.system("")

    parser = argparse.ArgumentParser(
        description='\033[1mConvert arknights story raw data into xlsx file.\033[m', epilog='\033[1mBy Nightsky#3319 in RIHQ: discord.gg/rihq\033[m')
    parser.add_argument('path', metavar='path', nargs='?', type=str, default='ArknightsGameData',
                        help='The file path or folder path. If no -E or -e option provided, convert all file in path or selected path to xlsx file')
    parser.add_argument('-c', '--comment', action='store_const', const=True,
                        default=False, help='Show Code Comment in raw story file')
    parser.add_argument('-L', '--Lang', metavar='Langcode', nargs=1, default=[
                        'zh_CN'], help='Config the language of the following command, default is zh_CN')
    parser.add_argument('-E', '--EventList', action='store_const', const=True, default=False,
                        help='Show available index: eventid in corresponding language, config the path to change the ArknightsGameData path from default (./ArknightsGameData)')
    parser.add_argument('-e', '--event', metavar='Eventid', nargs=1,
                        help='Export all stories in corresponding index or eventid, config the path to change the ArknightsGameData path from default (./ArknightsGameData). You can get available index or eventid from --EventList command.')
    parser.add_argument('-r', '--records', action='store_const', const=True,
                        default=False, help="Export all Operators' Records into one xlsx file.")
    parser.add_argument('-m', '--main', action='store_const',
                        const=True, default=False, help="Export all mainline story")
    parser.add_argument('-i', '--info', action='store_const',
                        const=True, default=False, help="Show story info in menu")
    args = parser.parse_args()
    #args=parser.parse_args(['-r','-i'])

    try:
        if args.EventList:
            actList = func.getAct(Path(args.path), args.Lang[0])
            mainList = func.getMainline(Path(args.path), args.Lang[0])
            recList = func.getRecords(Path(args. path), args.Lang[0])
            idx = 0
            print(
                f'\033[1mList of Events in {args.Lang[0]}:\033[m \n| Index: Eventid        Eventname')
            print('===ACTIVITY===')
            for event in actList:
                print(f'| {idx:<2}: {event.eventid:<20} {event.name:<}')
                idx += 1
            print('===MAINLINE===')
            for event in mainList:
                print(f'| {idx:<2}: {event.eventid:<20} {event.name:<}')
                idx += 1
            if len(recList):
                print('===RECORDS===')
                for event in recList:
                    print(f'| {idx:<2}: {event.eventid:<20} {event.name:<}')
                    idx += 1
            exit()
    except IndexError:
        pass

    commentFlag = args.comment
    infoFlag = args.info

    print("==========================")
    print("Status:")
    print("     Code Comment output: \033[{}m{}\033[m".format(
        '33;1' if commentFlag else '90', str(commentFlag)))
    print("     Story Info output: \033[{}m{}\033[m".format(
        '33;1' if infoFlag else '90', str(infoFlag)))
    print("==========================")

    if Path(args.path).is_dir():

        if args.main:  # Mainline stories

            records = func.getMainline(Path(args.path), args.Lang[0])

            filename = 'Mainline.xlsx'

            storyList = []
            txtList = []
            for event in records:
                for story in event:
                    storyList.append(story)
                    txtList.append(story.storyTxt)

            wb = xl.Workbook()
            ws = wb.active
            ws.title = 'Menu'
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 80

            ws.append([None, "Mainline"])
            ws['B1'].font = bold
            ws.append([])
            ws.append(['storyCode', 'storyName', 'avgTag',
                       'storyTxt.stem', 'storyInfo' if infoFlag else None])
            for cell in ws['A3':'E3'][0]:
                cell.font = bold

            for idx, story in enumerate(storyList):
                if infoFlag:
                    with open(story.storyInfo, encoding='utf-8') as storyInfoFile:
                        storyInfo = storyInfoFile.read()
                else:
                    storyInfo = None

                ws.append([story.storyCode, story.storyName,
                           story.avgTag, story.storyTxt.stem, storyInfo])
                loc = f"'{story.storyTxt.stem}'!A1"
                cell = ws.cell(idx+4, 4)
                cell.font = underline
                cell.hyperlink = Hyperlink(
                    ref='', location=loc, display=story.storyTxt.stem)
                for i in range(1, 5):
                    ws.cell(
                        idx+4, i).alignment = xl.styles.Alignment(vertical='center')
                if ws.cell(idx+3, 1).value == story.storyCode:
                    ws.merge_cells(start_column=1, end_column=1,
                                   start_row=idx+3, end_row=idx+4)
                    ws.merge_cells(start_column=2, end_column=2,
                                   start_row=idx+3, end_row=idx+4)

                for row in ws[ws.dimensions]:
                    row[4].alignment = xl.styles.Alignment(wrap_text=True)

        elif args.records:  # Operators' Records

            records = func.getRecords(Path(args.path), args.Lang[0])

            filename = 'Operators_Records.xlsx'

            storyList = []
            txtList = []
            for event in records:
                for story in event:
                    storyList.append(story)
                    txtList.append(story.storyTxt)

            wb = xl.Workbook()
            ws = wb.active
            ws.title = 'Menu'
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 80

            ws.append([None, "Operactors' Records"])
            ws['B1'].font = bold
            ws.append([])
            ws.append(['eventid', 'storyName', 'avgTag',
                       'storyTxt.stem', 'storyInfo' if infoFlag else None])
            for cell in ws['A3':'E3'][0]:
                cell.font = bold

            for idx, story in enumerate(storyList):
                if infoFlag:
                    with open(story.storyInfo, encoding='utf-8') as storyInfoFile:
                        storyInfo = storyInfoFile.read()
                else:
                    storyInfo = None

                ws.append([story.eventid, story.storyName,
                           story.avgTag, story.storyTxt.stem, storyInfo])
                loc = f"'{story.storyTxt.stem}'!A1"
                cell = ws.cell(idx+4, 4)
                cell.font = underline
                cell.hyperlink = Hyperlink(
                    ref='', location=loc, display=story.storyTxt.stem)

                for i in range(1, 5):
                    ws.cell(
                        idx+4, i).alignment = xl.styles.Alignment(vertical='center')
                if ws.cell(idx+3, 1).value == story.eventid:
                    ws.merge_cells(start_column=1, end_column=1,
                                   start_row=idx+3, end_row=idx+4)
                    ws.merge_cells(start_column=2, end_column=2,
                                   start_row=idx+3, end_row=idx+4)

                for row in ws[ws.dimensions]:
                    row[4].alignment = xl.styles.Alignment(wrap_text=True)

        elif args.event:  # Single Event
            try:
                eventid = list(func.getEvents(Path(args.path), args.Lang[0]))[
                    int(args.event[0])].eventid
            except ValueError:
                eventid = args.event[0]

            event = func.Event(Path(args.path), args.Lang[0], eventid)

            filename = f'{event.eventid}_{event.name}.xlsx'

            txtList = []
            for story in event:
                txtList.append(story.storyTxt)

            wb = xl.Workbook()
            ws = wb.active
            ws.title = 'Menu'
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 80

            ws.append([event.eventid, event.name])
            ws['B1'].font = bold
            ws.append([])
            ws.append(['storyCode', 'storyName', 'avgTag',
                       'storyTxt.stem', 'storyInfo' if infoFlag else None])
            for cell in ws['A3':'E3'][0]:
                cell.font = bold

            for idx, story in enumerate(event):
                if infoFlag:
                    with open(story.storyInfo, encoding='utf-8') as storyInfoFile:
                        storyInfo = storyInfoFile.read()
                else:
                    storyInfo = None

                ws.append([story.storyCode, story.storyName,
                           story.avgTag, story.storyTxt.stem, storyInfo])
                loc = f"'{story.storyTxt.stem}'!A1"
                cell = ws.cell(idx+4, 4)
                cell.font = underline
                cell.hyperlink = Hyperlink(
                    ref='', location=loc, display=story.storyTxt.stem)

                for i in range(1, 5):
                    ws.cell(
                        idx+4, i).alignment = xl.styles.Alignment(vertical='center')
                if ws.cell(idx+3, 1).value == story.storyCode and len(story.storyCode):
                    ws.merge_cells(start_column=1, end_column=1,
                                   start_row=idx+3, end_row=idx+4)
                    ws.merge_cells(start_column=2, end_column=2,
                                   start_row=idx+3, end_row=idx+4)

                for row in ws[ws.dimensions]:
                    row[4].alignment = xl.styles.Alignment(wrap_text=True)
        else:
            filename = 'story.xlsx'
            print("Target path type: \033[33;1mFolder\033[m")
            wb = xl.Workbook()
            txtList = []
            for txtFile in getFile(Path(args.path)):
                if txtFile.suffix == '.txt':
                    txtList.append(txtFile)

        print(f"{len(txtList)} txt files dectected")

        for (txtindex, txtFile) in enumerate(txtList):
            ws = wb.create_sheet(title=txtFile.stem)
            ws.column_dimensions['A'].hidden= True
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 70
            with open(txtFile,encoding='utf-8') as txtF:
                rawText = txtF.read()

            rawList = reader(rawText)
            xlc(ws,rawList)

            for row in ws[ws.dimensions]:
                for cell in row:
                    cell.alignment = xl.styles.Alignment(wrap_text=True)
            print(
                f"\rTxt file {txtFile.name} exported \033[1m({txtindex+1}/{len(txtList)})\033[m                 ")

        ws = wb.create_sheet(title='Characters')
        ws.append(['<Characters>'])
        for name in characters:
            ws.append([name])

        ws.append(['<Codes>'])
        for name in codes:
            ws.append([name])

        print("Character sheet exported")
        try:
            wb.save(filename=filename)
            print(f"\033[92mExported to \033[1m{filename}\033[m")
        except PermissionError:
            print('\033[91mPermissionError: Fail to save the file, maybe because you have already opened the file! Close the file and rerun.\033[m')
            sys.exit('PermissionError')

        
    else:
        print("Target path type: \033[33;1mFile\033[m")
        wb = xl.Workbook()
        ws = wb.active
        ws.title = Path(args.path).stem
        with open(Path(args.path),encoding='utf-8') as txtF:
            rawText = txtF.read()

        rawList = reader(rawText)
        xlc(ws,rawList)
        try:
            wb.save(filename=args.path.replace('.txt', '.xlsx'))
            print("\033[92mExported to \033[1m{}\033[m".format(
            args.path.replace('.txt', '.xlsx')))
        except PermissionError:
            print('\033[91mPermissionError: Fail to save the file, maybe because you have already opened the file! Close the file and rerun.\033[m')
            sys.exit('PermissionError')

